#!/usr/bin/env python3
"""Generate infra cost comparison Excel sheet for WhatsApp sharing."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colors ──────────────────────────────────────────
NAVY = "192B4D"
WHITE = "FFFFFF"
BLUE = "0E70B2"
LIGHT_BLUE = "E8F4FD"
GREEN = "16A34A"
GREEN_BG = "ECFDF5"
ORANGE = "EA580C"
ORANGE_BG = "FFF7ED"
GRAY_BG = "F7F9FB"
BORDER_COLOR = "DCE8F0"

thin_border = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)

header_font = Font(name="Arial", size=10, bold=True, color=WHITE)
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
title_font = Font(name="Arial", size=14, bold=True, color=NAVY)
subtitle_font = Font(name="Arial", size=9, color="7A8A9A")
data_font = Font(name="Arial", size=10, color="1E2A3A")
data_font_bold = Font(name="Arial", size=10, bold=True, color="1E2A3A")
blue_font = Font(name="Arial", size=10, bold=True, color=BLUE)
green_font = Font(name="Arial", size=10, bold=True, color=GREEN)
orange_font = Font(name="Arial", size=10, color=ORANGE)
muted_font = Font(name="Arial", size=9, color="7A8A9A")
rank_font = Font(name="Arial", size=11, bold=True, color=BLUE)

row_fill = PatternFill(start_color=GRAY_BG, end_color=GRAY_BG, fill_type="solid")
highlight_fill = PatternFill(
    start_color="E8F2FA", end_color="E8F2FA", fill_type="solid"
)
green_fill = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
right = Alignment(horizontal="right", vertical="center", wrap_text=True)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border


def style_data_row(ws, row, cols, zebra=False, highlight=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = thin_border
        cell.font = data_font
        if highlight:
            cell.fill = highlight_fill
        elif zebra:
            cell.fill = row_fill


# ═══════════════════════════════════════════════════
# Sheet 1: Combo Ranking
# ═══════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Infra Combos"
ws1.sheet_properties.tabColor = BLUE

# Column widths
widths = [5, 28, 22, 12, 12, 12, 11, 11, 9, 7]
for i, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Title
ws1.merge_cells("A1:J1")
ws1["A1"].value = "OCR + LLM Infrastructure Combos"
ws1["A1"].font = title_font
ws1["A1"].alignment = Alignment(horizontal="left", vertical="center")

ws1.merge_cells("A2:J2")
ws1["A2"].value = "35,000 pages/month · 420,000 pages/year · MYR at 1 USD = 4.40"
ws1["A2"].font = subtitle_font

# Headers
headers = [
    "#",
    "OCR Stack",
    "LLM",
    "Cost/mo\n(RM)",
    "Cost/page\n(RM)",
    "Max\npages/mo",
    "Reliability",
    "Quality",
    "Ops",
    "Score",
]
for c, h in enumerate(headers, 1):
    ws1.cell(row=4, column=c, value=h)
style_header(ws1, 4, len(headers))
ws1.row_dimensions[4].height = 32

# Data
combos = [
    (
        1,
        "PaddleOCR · Hetzner CX32",
        "Gemini 2.5 Flash",
        92,
        0.0026,
        "170K",
        "3/5",
        "4/5",
        "Self",
        4.1,
    ),
    (
        2,
        "PaddleOCR · Hetzner CX32",
        "DeepSeek V3 (OpenRouter)",
        73,
        0.0021,
        "170K",
        "3/5",
        "3.5/5",
        "Self",
        3.9,
    ),
    (
        3,
        "PaddleOCR · Hetzner CX32",
        "Gemini Flash Lite",
        62,
        0.0018,
        "170K",
        "3/5",
        "3/5",
        "Self",
        3.8,
    ),
    (
        4,
        "PaddleOCR · GCP e2-medium",
        "Gemini 2.5 Flash",
        167,
        0.0048,
        "170K",
        "4/5",
        "4/5",
        "Self",
        3.8,
    ),
    (
        5,
        "PaddleOCR · Hetzner CX32",
        "GPT-4o Mini (OpenRouter)",
        92,
        0.0026,
        "170K",
        "3.5/5",
        "3.5/5",
        "Self",
        3.8,
    ),
    (
        6,
        "PaddleOCR · Hetzner AX42",
        "Gemini 2.5 Flash",
        270,
        0.0077,
        "500K",
        "3.5/5",
        "4/5",
        "Self",
        3.8,
    ),
    (
        7,
        "Datalab Balanced API",
        "— (built-in)",
        616,
        0.0176,
        "10M+",
        "5/5",
        "4.5/5",
        "Zero",
        3.7,
    ),
    (
        8,
        "PaddleOCR · AWS t3a.medium",
        "GPT-4o Mini (OpenRouter)",
        179,
        0.0051,
        "170K",
        "4/5",
        "3.5/5",
        "Self",
        3.7,
    ),
    (
        9,
        "Datalab Balanced API",
        "Gemini Flash Lite",
        645,
        0.0184,
        "10M+",
        "5/5",
        "5/5",
        "Zero",
        3.6,
    ),
    (
        10,
        "Marker · RunPod A4000",
        "— (built-in)",
        634,
        0.0181,
        "1M+",
        "2.5/5",
        "4.5/5",
        "GPU",
        3.4,
    ),
    (
        11,
        "PaddleOCR · Hetzner CX32",
        "Claude Haiku 4.5",
        390,
        0.0111,
        "170K",
        "3/5",
        "4.5/5",
        "Self",
        3.4,
    ),
    (
        12,
        "PaddleOCR · RunPod A4000",
        "Gemini 2.5 Flash",
        693,
        0.0198,
        "1.3M",
        "2.5/5",
        "4/5",
        "GPU",
        3.2,
    ),
    (
        13,
        "Datalab High Accuracy",
        "— (built-in)",
        924,
        0.0264,
        "10M+",
        "5/5",
        "5/5",
        "Zero",
        3.1,
    ),
    (
        14,
        "Azure Doc Intelligence",
        "Gemini Flash Lite",
        1569,
        0.0448,
        "Unlim",
        "5/5",
        "4/5",
        "Min",
        2.4,
    ),
    (
        15,
        "AWS Textract",
        "Gemini Flash Lite",
        2339,
        0.0668,
        "Unlim",
        "5/5",
        "4/5",
        "Min",
        2.0,
    ),
]

for i, row_data in enumerate(combos):
    r = 5 + i
    ws1.row_dimensions[r].height = 22
    rank, ocr, llm, cost_mo, cost_pg, cap, rel, qual, ops, score = row_data

    ws1.cell(row=r, column=1, value=rank).alignment = center
    ws1.cell(row=r, column=2, value=ocr).alignment = left
    ws1.cell(row=r, column=3, value=llm).alignment = left
    ws1.cell(row=r, column=4, value=cost_mo).alignment = right
    ws1.cell(row=r, column=4).number_format = "#,##0"
    ws1.cell(row=r, column=5, value=cost_pg).alignment = right
    ws1.cell(row=r, column=5).number_format = "0.0000"
    ws1.cell(row=r, column=6, value=cap).alignment = center
    ws1.cell(row=r, column=7, value=rel).alignment = center
    ws1.cell(row=r, column=8, value=qual).alignment = center
    ws1.cell(row=r, column=9, value=ops).alignment = center
    ws1.cell(row=r, column=10, value=score).alignment = center

    is_highlight = rank in (1, 7)
    is_zebra = i % 2 == 1
    style_data_row(ws1, r, len(headers), zebra=is_zebra, highlight=is_highlight)

    # Bold the rank and score for top picks
    if rank in (1, 2, 3, 7):
        ws1.cell(row=r, column=1).font = rank_font
        ws1.cell(row=r, column=10).font = blue_font
    if rank == 1:
        ws1.cell(row=r, column=2).font = data_font_bold
        ws1.cell(row=r, column=4).font = green_font

# ═══════════════════════════════════════════════════
# Sheet 2: Margin Analysis
# ═══════════════════════════════════════════════════
ws2 = wb.create_sheet("Margin Analysis")
ws2.sheet_properties.tabColor = GREEN

widths2 = [5, 28, 22, 12, 12, 12, 10]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.merge_cells("A1:G1")
ws2["A1"].value = "Margin Analysis at RM 0.143/page (24-month plan)"
ws2["A1"].font = title_font

ws2.merge_cells("A2:G2")
ws2["A2"].value = "Revenue: RM 5,000/month · 35,000 pages/month"
ws2["A2"].font = subtitle_font

headers2 = [
    "#",
    "OCR Stack",
    "LLM",
    "Infra\nCost/mo",
    "Gross\nMargin",
    "Margin\n%",
    "Verdict",
]
for c, h in enumerate(headers2, 1):
    ws2.cell(row=4, column=c, value=h)
style_header(ws2, 4, len(headers2))
ws2.row_dimensions[4].height = 32

margins = [
    (
        1,
        "PaddleOCR · Hetzner CX32",
        "Gemini 2.5 Flash",
        92,
        4908,
        0.982,
        "Best overall",
    ),
    (3, "PaddleOCR · Hetzner CX32", "Gemini Flash Lite", 62, 4938, 0.988, "Cheapest"),
    (2, "PaddleOCR · Hetzner CX32", "DeepSeek V3", 73, 4927, 0.985, "Budget + quality"),
    (7, "Datalab Balanced API", "— (built-in)", 616, 4384, 0.877, "Zero ops"),
    (
        9,
        "Datalab Balanced",
        "Gemini Flash Lite",
        645,
        4355,
        0.871,
        "Best quality + zero ops",
    ),
    (13, "Datalab High Accuracy", "— (built-in)", 924, 4076, 0.815, "Premium quality"),
    (
        11,
        "PaddleOCR · Hetzner CX32",
        "Claude Haiku 4.5",
        390,
        4610,
        0.922,
        "Best LLM quality",
    ),
    (
        14,
        "Azure Doc Intelligence",
        "Gemini Flash Lite",
        1569,
        3431,
        0.686,
        "Overpriced",
    ),
    (15, "AWS Textract", "Gemini Flash Lite", 2339, 2661, 0.532, "Overpriced"),
]

for i, row_data in enumerate(margins):
    r = 5 + i
    ws2.row_dimensions[r].height = 22
    rank, ocr, llm, cost, margin, pct, verdict = row_data

    ws2.cell(row=r, column=1, value=rank).alignment = center
    ws2.cell(row=r, column=2, value=ocr).alignment = left
    ws2.cell(row=r, column=3, value=llm).alignment = left
    ws2.cell(row=r, column=4, value=cost).alignment = right
    ws2.cell(row=r, column=4).number_format = "#,##0"
    ws2.cell(row=r, column=5, value=margin).alignment = right
    ws2.cell(row=r, column=5).number_format = "#,##0"
    ws2.cell(row=r, column=6, value=pct).alignment = center
    ws2.cell(row=r, column=6).number_format = "0.0%"
    ws2.cell(row=r, column=7, value=verdict).alignment = center

    is_zebra = i % 2 == 1
    is_highlight = rank in (1, 7)
    style_data_row(ws2, r, len(headers2), zebra=is_zebra, highlight=is_highlight)

    if rank in (1, 3):
        ws2.cell(row=r, column=5).font = green_font
        ws2.cell(row=r, column=6).font = green_font
    if "Overpriced" in verdict:
        ws2.cell(row=r, column=7).font = orange_font

# Footer note
r_note = 5 + len(margins) + 1
ws2.merge_cells(f"A{r_note}:G{r_note}")
ws2[
    f"A{r_note}"
].value = "Note: Gross margin is before support staff (~RM 1,500-3,500/mo), maintenance, and admin overhead."
ws2[f"A{r_note}"].font = muted_font

# ═══════════════════════════════════════════════════
# Sheet 3: Quick Summary (WhatsApp-friendly)
# ═══════════════════════════════════════════════════
ws3 = wb.create_sheet("TL;DR")
ws3.sheet_properties.tabColor = "EA580C"

for i, w in enumerate([30, 14, 14, 14], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.merge_cells("A1:D1")
ws3["A1"].value = "Infra Cost Summary — Heng Hup OCR Service"
ws3["A1"].font = title_font

ws3.merge_cells("A2:D2")
ws3["A2"].value = "At 35K pages/month · selling at RM 0.143/page = RM 5K/mo revenue"
ws3["A2"].font = subtitle_font

headers3 = ["Stack", "Cost/mo (RM)", "Margin/mo (RM)", "Margin %"]
for c, h in enumerate(headers3, 1):
    ws3.cell(row=4, column=c, value=h)
style_header(ws3, 4, 4)

summary = [
    ("PaddleOCR + Gemini Flash (current)", 92, 4908, 0.982),
    ("PaddleOCR + DeepSeek V3 (cheapest)", 73, 4927, 0.985),
    ("PaddleOCR + Gemini Flash Lite", 62, 4938, 0.988),
    ("Datalab API (zero ops)", 616, 4384, 0.877),
    ("Datalab + Gemini (best quality)", 645, 4355, 0.871),
    ("Azure + Gemini (enterprise)", 1569, 3431, 0.686),
]

for i, (stack, cost, margin, pct) in enumerate(summary):
    r = 5 + i
    ws3.row_dimensions[r].height = 24
    ws3.cell(row=r, column=1, value=stack).alignment = left
    ws3.cell(row=r, column=1).font = data_font_bold if i < 2 else data_font
    ws3.cell(row=r, column=2, value=cost).alignment = right
    ws3.cell(row=r, column=2).number_format = "#,##0"
    ws3.cell(row=r, column=3, value=margin).alignment = right
    ws3.cell(row=r, column=3).number_format = "#,##0"
    ws3.cell(row=r, column=3).font = green_font if pct > 0.9 else data_font
    ws3.cell(row=r, column=4, value=pct).alignment = center
    ws3.cell(row=r, column=4).number_format = "0.0%"
    style_data_row(ws3, r, 4, zebra=(i % 2 == 1), highlight=(i == 0))

r_bottom = 5 + len(summary) + 1
ws3.merge_cells(f"A{r_bottom}:D{r_bottom}")
ws3[
    f"A{r_bottom}"
].value = "Infrastructure is <2% of revenue. The real cost is support & human time."
ws3[f"A{r_bottom}"].font = Font(name="Arial", size=10, bold=True, color=BLUE)

# ── Save ──
out = "/home/laughdiemeh/Projects/triple-h/proposal/heng-hup-infra-costs.xlsx"
wb.save(out)
print(f"Saved: {out}")
